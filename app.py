import os
import io
import re
import time
import json
import base64
import math
import statistics
from PIL import Image
from flask import Flask, request, jsonify, render_template, send_file
from dotenv import load_dotenv
from openai import OpenAI
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
app = Flask(__name__)
MODELO = "gpt-4o"           # Melhor para leitura visual de layouts
# MODELO = "gpt-4.1"        # Alternativa nova da OpenAI


VARIANTES = {
    "DERMO": ["DERMO", "DEMO", "PERFUMARIA"],
    "ESMALTES": ["ESMALTES", "ESMALTE", "ESM"],
    "PF CANALETADO": ["PF CANALETADO", "CANALETADO", "PAINEL CANALETADO", "PF CANAL 807", "PF CANALETADO 807"],
    "PF 807": ["PF 807", "PF 807MM"],
    "PF 807 FUNDO": ["PF 807 FUNDO", "PF 807MM FUNDO"],
    "PF 550": ["PF 550", "PF 550MM"],
    "PF 1000": ["PF 1000", "PF 1000MM"],
    "GOND 3000": ["GOND 3000", "GOND 3000MM", "GONDOLA 3000"],
    "GOND 2200": ["GOND 2200", "GOND 2200MM", "GONDOLA 2200"],
    "GOND 2000": ["GOND 2000", "GOND 2000MM", "GONDOLA 2000"],
    "GOND 1700": ["GOND 1700", "GOND 1700MM", "GONDOLA 1700"],
    "GOND 1400": ["GOND 1400", "GOND 1400MM", "GONDOLA 1400"],
    "GOND": ["GOND", "GONDOLA"],
    "MIP 500": ["MIP 500", "MIP 500MM"],
    "LAT CX 400": ["LAT CX 400", "LAT. CAIXA 400"],
    "LAT CX 550": ["LAT CX 550", "LAT. CAIXA 550", "LAT CAIXA 550", "LAT CAIXA", "LATERAL CAIXA", "CAIXA 550"],
    "MAQ 500": ["MAQ 500", "MAQ 500MM"],
    "BA 800": ["BA 800", "BA 800MM", "PDV 800", "PDV 800MM", "BALCAO 800", "BALCÃO 800"],
    "BA 700": ["BA 700", "BA 700MM", "PDV 700", "PDV 700MM", "BALCAO 700", "BALCÃO 700"],
    "BA 600": ["BA 600", "BA 600MM", "PDV 600", "PDV 600MM", "BALCAO 600", "BALCÃO 600"],
    "BA 1000": ["BA 1000", "BA 1000MM", "PDV 1000", "PDV 1000MM", "BALCAO 1000", "BALCÃO 1000", "BALCAO", "BALCÃO"],
    "BA VIDRO 1000": ["BA VIDRO 1000", "BA VIDRO 1000MM"],
    "BA VIDRO 800": ["BA VIDRO 800", "BA VIDRO 800MM"],
    "BA VIDRO 700": ["BA VIDRO 700", "BA VIDRO 700MM"],
    "BA VIDRO 600": ["BA VIDRO 600", "BA VIDRO 600MM"],
    "BA PIA 900": ["BA PIA 900", "BA PIA 900MM"],
    "BA POMBAL 1000": ["BA POMBAL 1000", "BA POMBAL", "POMBAL 1000", "POMBAL"],
    "CESTAO": ["CESTAO", "CESTÃO", "CESTÃO 400", "CESTAO 400", "CESTO", "CEST 400", "CESTO 400"],
    "CHECKOUT 1000": ["CHECKOUT 1000", "CHECK OUT 1000"],
    "CAIXA 1000": ["CAIXA 1000", "CAIXA 1000MM"],
    "CAIXA 600": ["CAIXA 600", "CHECKOUT 600", "CHECK OUT 600"],
    "CHECKOUT L": ["CHECKOUT L"],
    "VITRINE": ["VITRINE", "VITRINE 807", "VITRINE 807MM"],
    "MED 807": ["MED 807", "MED 807MM"],
    "MED 500": ["MED 500", "MED 500MM"],
    "CONTROLADO": ["CONTROLADO", "MED CONTROLADO", "CTRL 500"],
    "CANTONEIRA 400": ["CANTONEIRA 400", "CANTONEIRA 400MM"],
    "PORTA CORRER": ["PORTA CORRER"],
    "PORTA VAI VEM": ["PORTA VAI VEM"],
    "FECHAMENTO": ["FECHAMENTO"],
    "BASE 1200": ["BASE 1200"],
    "MESA EM L": ["MESA EM L", "MESA L", "MESA EM L AMADEIRADO"],
    "ESPACO KIDS": ["ESPACO KIDS", "ESPAÇO KIDS", "KIDS"],
    "BOMB": ["BOMB", "BOMBA"],
    "MACA": ["MACA", "MACA/ESCADA"],
}
LISTA_OFFICIAL = ", ".join(VARIANTES.keys())

# Ordem de prioridade: mais específico primeiro
ORDEM_EXTRACAO = [
    "BA POMBAL 1000", "BA VIDRO 1000", "BA VIDRO 800", "BA VIDRO 700", "BA VIDRO 600",
    "BA PIA 900", "BA 1000", "BA 800", "BA 700", "BA 600",
    "LAT CX 550", "LAT CX 400",
    "GOND 3000", "GOND 2200", "GOND 2000", "GOND 1700", "GOND 1400",
    "CHECKOUT 1000", "CAIXA 1000", "CAIXA 600", "CHECKOUT L",
    "PF CANALETADO", "PF 807 FUNDO", "PF 807", "PF 550", "PF 1000",
    "MED 807", "MED 500", "CONTROLADO",
    "MIP 500", "MAQ 500", "VITRINE", "CESTAO", "ESMALTES", "DERMO",
    "CANTONEIRA 400", "PORTA CORRER", "PORTA VAI VEM",
    "FECHAMENTO", "BASE 1200", "MESA EM L", "ESPACO KIDS", "BOMB", "MACA",
    "GOND",
]

def extrair_id(texto):
    if not texto: return None
    t = str(texto).upper().strip()
    for oficial, kws in VARIANTES.items():
        for kw in kws:
            if kw in t: return oficial
    return None

def pil_to_b64(img):
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()

def analisar_recorte_ia(b64, prompt_idx, custom_prompt=None):
    """Analisa um recorte e retorna caixas delimitadoras (x, y, w, h)."""
    base_prompt = """VOCE EH UM ESPECIALISTA EM INVENTARIO.
Conte os moveis desta LISTA: {lista}.

PARA CADA MOVEL:
1. Identifique o retangulo exato do movel.
2. Leia o texto interno (ex: "MED 807mm").
3. Retorne a caixa delimitadora (x, y, largura, altura) em escala 0-1000.

Retorne APENAS JSON:
{{"itens": [{{"n": "NOME", "x": 100, "y": 100, "w": 50, "h": 200}}]}}"""
    
    prompt = base_prompt.replace("{lista}", str(LISTA_OFFICIAL))
        
    try:
        resp = client.chat.completions.create(
            model=MODELO,
            temperature=0.0,
            seed=42,
            response_format={"type": "json_object"},
            messages=[{"role": "user", "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}", "detail": "high"}}
            ]}]
        )
        content = resp.choices[0].message.content or "{}"
        try:
            data = json.loads(content)
            raw_itens = data.get("itens", [])
            validos = []
            for it in raw_itens:
                nome = extrair_id(it.get("n", ""))
                if nome:
                    validos.append({
                        "nome": nome,
                        "x": it.get("x", 0),
                        "y": it.get("y", 0),
                        "w": it.get("w", 20),
                        "h": it.get("h", 20)
                    })
            return validos, content
        except json.JSONDecodeError:
            return [], content
    except Exception as e:
        return [], str(e)

def analisar_recorte(img, rotacoes, crop_x, crop_y):
    """Processa recorte com caixas delimitadoras."""
    ang = rotacoes[0] if rotacoes else 0
    rot = img if ang == 0 else img.rotate(ang, expand=True)
    w_px, h_px = rot.size
    b64 = pil_to_b64(rot)

    all_raw = []
    textos = []
    with ThreadPoolExecutor(max_workers=2) as executor:
        futures = [executor.submit(analisar_recorte_ia, b64, 0) for _ in range(2)]
        for f in futures:
            items, txt = f.result()
            all_raw.extend(items)
            textos.append(txt)

    itens_globais = []
    for it in all_raw:
        # Converte 0-1000 -> Pixels -> PDF Units
        gx = crop_x + (it["x"] / 1000.0) * w_px
        gy = crop_y + (it["y"] / 1000.0) * h_px
        gw = (it["w"] / 1000.0) * w_px
        gh = (it["h"] / 1000.0) * h_px
        
        itens_globais.append({
            "nome": it["nome"],
            "x1": gx, "y1": gy,
            "x2": gx + gw, "y2": gy + gh
        })

    os.makedirs("debug", exist_ok=True)
    img_path = f"debug/crop_{int(time.time()*1000)}.png"
    rot.save(img_path)
    return itens_globais, textos[0], img_path

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/renderizar-pdf", methods=["POST"])
def renderizar_pdf():
    """Renderiza uma pagina do PDF em baixa resolucao para exibir no canvas."""
    pdf_file = request.files.get("pdf")
    pagina = int(request.form.get("pagina", 1))
    if not pdf_file:
        return jsonify({"erro": "Nenhum PDF enviado"}), 400
    pdf_bytes = pdf_file.read()
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    if pagina < 1 or pagina > len(doc):
        return jsonify({"erro": f"Pagina {pagina} invalida. PDF tem {len(doc)} paginas."}), 400
    page = doc[pagina - 1]
    # Baixa resolucao para o canvas (rapido)
    pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    b64 = pil_to_b64(img)
    return jsonify({"imagem": b64, "largura": pix.width, "altura": pix.height, "paginas": len(doc)})

@app.route("/debug_img/<filename>")
def debug_img(filename):
    return send_file(os.path.join("debug", filename))

def calcular_iou(boxA, boxB):
    # Determina as coordenadas da interseção
    xA = max(boxA[0], boxB[0])
    yA = max(boxA[1], boxB[1])
    xB = min(boxA[2], boxB[2])
    yB = min(boxA[3], boxB[3])
    
    interWidth = max(0, xB - xA)
    interHeight = max(0, yB - yA)
    interArea = interWidth * interHeight
    
    # Calcula a área de ambos os boxes
    boxAArea = (boxA[2] - boxA[0]) * (boxA[3] - boxA[1])
    boxBArea = (boxB[2] - boxB[0]) * (boxB[3] - boxB[1])
    
    if boxAArea + boxBArea - interArea == 0: return 0
    return interArea / float(boxAArea + boxBArea - interArea)

@app.route("/analisar/canvas", methods=["POST"])
def analisar_canvas():
    try:
        pdf_file = request.files.get("pdf")
        pagina = int(request.form.get("pagina", 1))
        recortes_json = request.form.get("recortes", "[]")
        if not pdf_file:
            return jsonify({"erro": "Nenhum PDF enviado"}), 400

        recortes = json.loads(recortes_json)
        pdf_bytes = pdf_file.read()
        ZOOM_IA = 18.0
        scale_x = scale_y = 2.0

        todos_itens_pag = []
        logs = []
        
        for i, r in enumerate(recortes):
            doc_local = fitz.open(stream=pdf_bytes, filetype="pdf")
            page_local = doc_local[pagina - 1]
            x1, y1, w, h = r["x"], r["y"], r["width"], r["height"]
            x1_pdf, y1_pdf = x1 / scale_x, y1 / scale_y
            x2_pdf, y2_pdf = (x1 + w) / scale_x, (y1 + h) / scale_y
            rect = fitz.Rect(x1_pdf, y1_pdf, x2_pdf, y2_pdf)

            pix = page_local.get_pixmap(matrix=fitz.Matrix(ZOOM_IA, ZOOM_IA), clip=rect)
            crop_img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            doc_local.close()
            
            resultado_itens, texto_ia, path_img = analisar_recorte(crop_img, r.get("rotacoes", [0]), x1_pdf, y1_pdf)
            todos_itens_pag.extend(resultado_itens)
            logs.append({"recorte": i+1, "texto": texto_ia, "imagem": path_img})

        # --- DEDUPLICAÇÃO POR IoU ---
        itens_finais = Counter()
        ja_processados = [False] * len(todos_itens_pag)
        IOU_THRESHOLD = 0.3 # Se sobrepor 30%, eh o mesmo movel

        for i in range(len(todos_itens_pag)):
            if ja_processados[i]: continue
            it1 = todos_itens_pag[i]
            itens_finais[it1["nome"]] += 1
            ja_processados[i] = True
            
            box1 = (it1["x1"], it1["y1"], it1["x2"], it1["y2"])
            
            for j in range(i + 1, len(todos_itens_pag)):
                if ja_processados[j]: continue
                it2 = todos_itens_pag[j]
                if it1["nome"] == it2["nome"]:
                    box2 = (it2["x1"], it2["y1"], it2["x2"], it2["y2"])
                    iou = calcular_iou(box1, box2)
                    if iou > IOU_THRESHOLD:
                        ja_processados[j] = True

        inventario = [{"nome": k, "quantidade": v} for k, v in sorted(itens_finais.items())]
        return jsonify({"inventario": inventario, "total": sum(itens_finais.values()), "logs": logs})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"erro": str(e)}), 500

@app.route("/exportar/excel", methods=["POST"])
def exportar_excel():
    """Gera um arquivo Excel formatado com o inventario."""
    data = request.get_json()
    inventario = data.get("inventario", [])
    nome_cliente = data.get("cliente", "Proposta")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventário"

    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_align = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.row_dimensions[1].height = 30

    ws["A1"] = "PRODUTO"
    ws["B1"] = "QUANTIDADE"
    for cell in [ws["A1"], ws["B1"]]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Linhas de dados
    alt_fill = PatternFill(start_color="f0f4ff", end_color="f0f4ff", fill_type="solid")
    for i, item in enumerate(inventario, start=2):
        ws[f"A{i}"] = item["nome"]
        ws[f"B{i}"] = item["quantidade"]
        ws[f"B{i}"].alignment = Alignment(horizontal="center")
        if i % 2 == 0:
            ws[f"A{i}"].fill = alt_fill
            ws[f"B{i}"].fill = alt_fill

    # Total
    total_row = len(inventario) + 3
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"B{total_row}"] = sum(item["quantidade"] for item in inventario)
    for cell in [ws[f"A{total_row}"], ws[f"B{total_row}"]]:
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="00d4ff", end_color="00d4ff", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"inventario_{nome_cliente.replace(' ', '_')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/analisar/recortes", methods=["POST"])
def analisar_recortes():
    """Modo legado: upload de imagens pre-cortadas."""
    files = request.files.getlist("imagens")
    if not files:
        return jsonify({"erro": "Nenhuma imagem enviada"}), 400
    contagem = Counter()
    for f in files:
        img = Image.open(f.stream).convert("RGB")
        resultado = analisar_com_consenso_rotacao(img, [0, 90, 180, 270])
        for nome, qtd in resultado.items():
            contagem[nome] += qtd
    inventario = [{"nome": k, "quantidade": v} for k, v in sorted(contagem.items())]
    return jsonify({"inventario": inventario, "total": sum(contagem.values())})

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
